import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

db_name = "EWP"
output_path = (
    r"C:\Users\I522148\OneDrive - SAP SE\SWAT\PLAs\Q3-H1"
    r"\Bocar Servicios SA de CV"
    r"\HANA_health_check_summary_observations_EWP.xlsx"
)

rows_data = [
    (
        "Parameters Out of ECS Standard",
        "Medium",
        (
            "A total of 9 parameters were found to deviate from the ECS standard configuration. Critical deviations "
            "include: daemon.ini [nameserver/environment] contains non-standard runtime values (PYTHONDONTWRITEBYTECODE=1, "
            "HDB_SINGLE_THREAD_GROUP=1); global.ini [inifile_checker/replicate] is set to 'false' instead of required "
            "'true', preventing INI file replication to secondary nodes; global.ini [sql/hex_enable_distributed_query_"
            "processing] is explicitly 'false' instead of the expected empty default; indexserver.ini [authorization/"
            "blocked_system_privileges_for_sys_repo_user] is missing the required privilege restriction list, a security "
            "compliance gap; indexserver.ini [lobhandling/garbage_collect_interval_s] is set to 0, disabling LOB garbage "
            "collection; indexserver.ini [mergedog/max_cpuload_for_merge] is set to 100, removing the CPU cap for merge "
            "operations; and two additional parameters in the calcengine and cache sections require remediation. All "
            "deviations should be corrected in coordination with the customer to align with ECS operational and security standards."
        )
    ),
    (
        "P1 Alerts - Last 40 Days",
        "High",
        (
            "A total of 58 Priority 1 alerts were triggered on tenant EWP over the last 40 days. Alert 140 (40 occurrences) "
            "indicates that the production system license usage type is not configured, a direct compliance and licensing "
            "violation requiring immediate resolution by the customer. Alert 39 (18 occurrences) reports multiple "
            "long-running SQL statements executing between 1,800 and 2,692 seconds across different sessions, pointing to "
            "workload management deficiencies including potential missing indexes, inefficient query execution plans, or "
            "resource contention under peak load. Recommended immediate actions: set the production license type in the "
            "HANA system configuration and engage the customer application team to investigate and tune the recurring "
            "long-running SQL workloads."
        )
    ),
    (
        "Large Technical Tables (Memory and Disk Footprint)",
        "Medium",
        (
            "31 large technical tables were detected in tenant EWP, consuming a combined 294.28 GB of memory and 1,386.65 "
            "GB on disk. The top contributors are: SOFFCONT1 (787.78 GB disk / 42.75 GB memory - document content), "
            "DBTABLOG (66.43 GB disk / 66.42 GB memory - change document log), SWWCNTP0 (34.62 GB memory - workflow "
            "counters), APQD (19.50 GB memory - async RFC queue), and CDPOS (19.13 GB memory - change document positions). "
            "Several tables including SOFFCONT1, ODQDATA_F, and GRACSODREPDATA exhibit disk usage significantly exceeding "
            "their memory footprint, indicating large volumes of cold data or LOB content. Recommended actions: implement "
            "archiving for SOFFCONT1 document content, establish a retention policy for DBTABLOG, and evaluate ODQ tables "
            "for data cleanup to reduce disk and memory consumption."
        )
    ),
    (
        "Out of Memory (OOM) Events",
        "High",
        (
            "One Out of Memory event was recorded on tenant EWP during the review period, occurring on July 30, 2026 at "
            "10:11 (indexserver compositelimit_oom trace file). The event was triggered by the indexserver exceeding its "
            "composite memory limit. Current memory utilization stands at approximately 1,575 GB in use out of 2,080 GB "
            "allocated to the indexserver, with total physical RAM of 3,831 GB. The primary memory consumers are the "
            "Column Store dictionary (349 GB), uncompressed main storage (237 GB), and index single columns (208 GB). "
            "Recommended actions: review and adjust the composite memory limit, investigate the workload that triggered "
            "the event, and evaluate data tiering or archiving options to reduce the overall memory footprint."
        )
    ),
    (
        "Top 10 Tables with Most Transactions (Record Growth - Last 40 Days)",
        "Medium",
        (
            "Based on record growth history from July 6 to August 17, 2026, the top 10 most transactionally active tables "
            "are: (1) MLDOC +26.7M records (+5.55%), (2) EDID4 +19.6M records (+3.45%), (3) GLPCA +16.1M records (+1.52%), "
            "(4) MLDOCCCS +12.4M records (+4.96%), (5) BNK_XSTAT_TX +12.3M records (+2.12%), (6) BNK_BTCH_STREF +12.3M "
            "records (+2.12%), and (7-10) ACDOCA across 4 partitions each adding approximately 9.1-9.2M records (+1.40%). "
            "The ACDOCA Universal Journal collectively accounts for over 36.6M new records, making it the highest-volume "
            "transactional entity in the system. MLDOC and MLDOCCCS exhibit above-average growth rates, indicating intense "
            "material ledger activity. These tables should be monitored for delta merge efficiency, compression ratios, "
            "and potential impact on memory and I/O performance."
        )
    ),
    (
        "Plan Cache Hit Ratio = 0%",
        "High",
        (
            "The SQL Plan Cache hit ratio on tenant EWP over the last 40 days is 0%, meaning no query execution plans "
            "are being reused from cache. Every statement triggers a full plan compilation, significantly increasing CPU "
            "consumption and execution latency across the system. This is a direct contributing factor to the long-running "
            "SQL statements reported under P1 Alerts and represents a systemic performance risk under concurrent workloads. "
            "Recommended actions: investigate plan cache sizing configuration (indexserver.ini [sql/plan_cache_size]), "
            "review workload patterns causing excessive cache invalidation or eviction, and analyze whether statement "
            "parameterization can be improved at the application layer to improve plan reuse."
        )
    ),
    (
        "Row Store Fragmentation",
        "Medium",
        (
            "Internal row store fragmentation is at 6.70% and external fragmentation at 2.21%. While not yet at critical "
            "levels, internal fragmentation of this magnitude indicates wasted memory within row store pages and can "
            "progressively degrade query performance and increase memory consumption for row store objects. This is "
            "particularly relevant given the system's existing memory pressure. Recommended actions: schedule a row store "
            "reorganization (ALTER SYSTEM RECLAIM VERSION SPACE) during a planned maintenance window, review which "
            "large row store tables are most fragmented, and evaluate candidates for migration to column store where the "
            "access pattern supports it."
        )
    ),
    (
        "LOB Garbage Collection Deactivated (Alert 98 - Continuously Firing)",
        "High",
        (
            "Alert 98 has been firing every hour continuously since at least August 16, 2026, confirming that LOB garbage "
            "collection is deactivated on the system. This is caused by indexserver.ini [lobhandling/garbage_collect_interval_s] "
            "being set to 0. As a result, deleted or orphaned LOB data is never reclaimed from disk, causing the LOB "
            "footprint to grow indefinitely. Given that SOFFCONT1 alone occupies 787 GB on disk with a high LOB content "
            "ratio, and combined with the data volume already near capacity (see Data Volume Risk), this represents a "
            "critical housekeeping gap with direct disk exhaustion risk. Immediate actions required: correct the parameter "
            "to re-enable LOB GC, trigger a manual LOB garbage collection cycle, and monitor LOB disk space reclamation."
        )
    ),
    (
        "Lock Wait Contention on Key Business Tables",
        "High",
        (
            "Significant transactional lock wait contention was identified on multiple business-critical tables. "
            "ZTPY_FORTIA_DISP accumulated 31,495 lock waits totaling 320,808 seconds of combined wait time (avg 10.18s). "
            "BNK_BATCH_HEADER recorded 71,940 waits totaling 180,214 seconds (avg 2.50s). SALV_CSQ_PARAMS shows a "
            "critically high average wait of 165 seconds across 33 events with 1 lock failure, suggesting a serialization "
            "design flaw. TSPEVJOB recorded an average wait of 451 seconds per event with 3 lock failures, indicating "
            "job scheduling table contention. NRIV accumulated 18,985 lock waits, a classic indicator of number range "
            "buffer undersizing. Recommended actions: review application locking logic for ZTPY_FORTIA_DISP and "
            "BNK_BATCH_HEADER, increase NRIV number range buffer sizes, and engage the Basis team to resolve TSPEVJOB "
            "and SALV_CSQ_PARAMS contention patterns."
        )
    ),
    (
        "Transaction Deadlocks (2 Events Detected)",
        "Medium",
        (
            "Two transaction deadlock events (Alert 149, MEDIUM severity) were recorded during the review period: "
            "August 13, 2026 at 23:55 and July 30, 2026 at 17:55. While the frequency is low, deadlocks indicate "
            "circular lock dependencies between concurrent transactions that result in one transaction being silently "
            "rolled back by the database engine. This can cause undetected data inconsistencies or silent business "
            "process failures if the application does not properly handle the rollback and retry. Recommended actions: "
            "identify the ABAP programs or background jobs running at the time of each deadlock event using the HANA "
            "trace files, analyze the locking sequence, and redesign the conflicting access patterns to eliminate the "
            "circular dependency."
        )
    ),
    (
        "System Replication Running in ASYNC Mode",
        "Medium",
        (
            "HANA System Replication (HSR) is active from the primary host vhbozehpdb01 to secondary hec48v147239 in "
            "ASYNC mode with logreplay operation. While the replication is currently ACTIVE with zero shipping delay, "
            "ASYNC mode carries an inherent risk of data loss in a failover scenario: transactions confirmed to the "
            "application may not yet have been shipped to or replayed on the secondary at the moment of a primary "
            "failure. For a productive ERP system of this scale, this RPO exposure may conflict with business continuity "
            "requirements. Recommended actions: review the agreed RTO and RPO targets with the customer, assess whether "
            "SYNCMEM replication mode is feasible given network latency between primary and secondary sites, and "
            "document the accepted data loss risk if ASYNC is intentionally maintained."
        )
    ),
]

ORANGE_COLOR = "FFA500"
HEADER_BLUE = "4472C4"
WHITE_COLOR = "FFFFFF"

num_cols = 8
last_data_row = 2 + len(rows_data)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "HANA Health Check Analysis"

ws.merge_cells("A1:" + get_column_letter(num_cols) + "1")
cap = ws["A1"]
cap.value = "DB Health check summary"
cap.font = Font(bold=True, size=14)
cap.alignment = Alignment(horizontal="center", vertical="center")
cap.fill = PatternFill(start_color=ORANGE_COLOR, end_color=ORANGE_COLOR, fill_type="solid")
ws.row_dimensions[1].height = 35

headers = [
    "Item Num", "Type", "Priority", "Item",
    "Observations", "System", "Responsible", "Link to section"
]

for col_idx, hdr in enumerate(headers, start=1):
    cell = ws.cell(row=2, column=col_idx, value=hdr)
    cell.font = Font(bold=True, color=WHITE_COLOR)
    cell.fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 30

HIGH_FILL = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
MED_FILL = PatternFill(start_color="FFFAE0", end_color="FFFAE0", fill_type="solid")

for i, (item_text, priority, obs_text) in enumerate(rows_data):
    row_num = i + 3
    row_fill = HIGH_FILL if priority == "High" else MED_FILL

    ws.cell(row=row_num, column=1, value=i + 1)
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    ws.cell(row=row_num, column=1).fill = row_fill

    ws.cell(row=row_num, column=2, value="Database")
    ws.cell(row=row_num, column=2).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    ws.cell(row=row_num, column=2).fill = row_fill

    ws.cell(row=row_num, column=3, value=priority)
    ws.cell(row=row_num, column=3).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    ws.cell(row=row_num, column=3).fill = row_fill

    ws.cell(row=row_num, column=4, value=item_text)
    ws.cell(row=row_num, column=4).alignment = Alignment(vertical="top", wrap_text=True)
    ws.cell(row=row_num, column=4).fill = row_fill

    ws.cell(row=row_num, column=5, value=obs_text)
    ws.cell(row=row_num, column=5).alignment = Alignment(vertical="top", wrap_text=True)
    ws.cell(row=row_num, column=5).fill = row_fill

    ws.cell(row=row_num, column=6, value=db_name)
    ws.cell(row=row_num, column=6).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    ws.cell(row=row_num, column=6).fill = row_fill

    ws.cell(row=row_num, column=7, value="Customer / TSM")
    ws.cell(row=row_num, column=7).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    ws.cell(row=row_num, column=7).fill = row_fill

    ws.cell(row=row_num, column=8, value="")
    ws.cell(row=row_num, column=8).alignment = Alignment(vertical="top", wrap_text=True)
    ws.cell(row=row_num, column=8).fill = row_fill

    ws.row_dimensions[row_num].height = 130

ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 50
ws.column_dimensions["E"].width = 100
ws.column_dimensions["F"].width = 12
ws.column_dimensions["G"].width = 18
ws.column_dimensions["H"].width = 18

thick_side = Side(style="thick")
thin_side = Side(style="thin")

for row_idx in range(2, last_data_row + 1):
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        top = thick_side if row_idx == 2 else thin_side
        bottom = thick_side if row_idx == last_data_row else thin_side
        left = thick_side if col_idx == 1 else thin_side
        right = thick_side if col_idx == num_cols else thin_side
        cell.border = Border(top=top, bottom=bottom, left=left, right=right)

wb.save(output_path)
print("Saved:", output_path)
