import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

db_name = "EWP"
output_path = (
    r"C:\Users\I522148\OneDrive - SAP SE\SWAT\PLAs\Q3-H1"
    r"\Bocar Servicios SA de CV"
    r"\HANA_health_check_summary_observations_EWP_MainSections.xlsx"
)

rows_data = [
    (
        "HANA DB Installed Version",
        (
            "The HANA database is running version 2.00.089.00 (SPS08 Rev 089.00), which is the latest recommended "
            "revision. No upgrade is currently required. The system is fully up to date with the most recent SAP HANA "
            "SPS08 patch level, ensuring stability, security patch coverage, and full SAP support eligibility. Ongoing "
            "monitoring for future patch releases is recommended to maintain compliance with ECS versioning standards."
        )
    ),
    (
        "Parameters Out of ECS Standard",
        (
            "A total of 9 parameters were found to deviate from the ECS standard configuration. Critical deviations "
            "include: daemon.ini [nameserver/environment] contains non-standard runtime values (PYTHONDONTWRITEBYTECODE=1, "
            "HDB_SINGLE_THREAD_GROUP=1); global.ini [inifile_checker/replicate] is set to 'false' instead of required "
            "'true', preventing INI file replication to secondary nodes; global.ini [sql/hex_enable_distributed_query_"
            "processing] is explicitly 'false' instead of the expected empty default; indexserver.ini [authorization/"
            "blocked_system_privileges_for_sys_repo_user] is missing the required privilege restriction list, representing "
            "a security compliance gap; indexserver.ini [lobhandling/garbage_collect_interval_s] is set to 0, potentially "
            "disabling LOB garbage collection; indexserver.ini [mergedog/max_cpuload_for_merge] is set to 100, removing "
            "the intended CPU cap for merge operations; and two additional parameters in the calcengine and cache sections "
            "also require remediation. All deviations should be corrected in coordination with the customer to align with "
            "ECS operational and security standards."
        )
    ),
    (
        "P1 Alerts - Last 40 Days",
        (
            "A total of 58 Priority 1 alerts were triggered on tenant EWP over the last 40 days. Alert 140 (40 occurrences) "
            "indicates that the production system license usage type is not configured, representing a direct compliance and "
            "licensing violation that must be resolved immediately by the customer. Alert 39 (18 occurrences) reports "
            "multiple long-running SQL statements executing between 1,800 and 2,692 seconds across different sessions, "
            "pointing to workload management deficiencies including potential missing indexes, inefficient query execution "
            "plans, or resource contention under peak load. Recommended immediate actions include setting the production "
            "license type in the HANA system configuration and engaging the customer application team to investigate and "
            "tune the recurring long-running SQL workloads."
        )
    ),
    (
        "Large Technical Tables (Memory and Disk Footprint)",
        (
            "31 large technical tables were detected in tenant EWP, consuming a combined 294.28 GB of memory and 1,386.65 "
            "GB on disk. The top contributors are: SOFFCONT1 (787.78 GB disk / 42.75 GB memory - document content), "
            "DBTABLOG (66.43 GB disk / 66.42 GB memory - change document log), SWWCNTP0 (34.62 GB memory - workflow "
            "counters), APQD (19.50 GB memory - async RFC queue), and CDPOS (19.13 GB memory - change document positions). "
            "Several tables including SOFFCONT1, ODQDATA_F, and GRACSODREPDATA exhibit disk usage significantly exceeding "
            "their memory footprint, indicating large volumes of cold data or LOB content. Recommended actions include "
            "implementing archiving for SOFFCONT1 document content, establishing a retention policy for DBTABLOG, and "
            "evaluating ODQ tables for data cleanup to reduce disk and memory consumption."
        )
    ),
    (
        "User Tables or Partitions Bigger Than 1.5 Billion Records",
        (
            "No user tables or individual partitions exceeding 1.5 billion records were identified in either the SystemDB "
            "or tenant EWP. Current table sizes are within acceptable thresholds and do not present immediate risks "
            "associated with partition key range exhaustion or extreme single-table memory pressure. However, several "
            "large business tables are growing steadily including GLPCA (1.075 billion records), BSEG (970 million), "
            "ACCTCR (753 million), and ACDOCA (approximately 662 million per partition). Proactive monitoring and "
            "archiving strategies for these tables are recommended to prevent future threshold breaches."
        )
    ),
    (
        "Out of Memory (OOM) Events",
        (
            "One Out of Memory event was recorded on tenant EWP during the review period, occurring on July 30, 2026 at "
            "10:11 (indexserver compositelimit_oom trace file). The event was triggered by the indexserver exceeding its "
            "composite memory limit. Current memory utilization stands at approximately 1,575 GB in use out of 2,080 GB "
            "allocated to the indexserver, with total physical RAM of 3,831 GB. The primary memory consumers are the "
            "Column Store dictionary (349 GB), uncompressed main storage (237 GB), and index single columns (208 GB). "
            "Recommended actions include reviewing and adjusting the composite memory limit, investigating the workload "
            "that triggered the event, and evaluating data tiering or archiving options to reduce the overall memory footprint."
        )
    ),
    (
        "Top 10 Tables with Most Transactions (Record Growth - Last 40 Days)",
        (
            "Based on record growth history from July 6 to August 17, 2026, the top 10 most transactionally active tables "
            "are: (1) MLDOC +26.7M records (+5.55%), (2) EDID4 +19.6M records (+3.45%), (3) GLPCA +16.1M records (+1.52%), "
            "(4) MLDOCCCS +12.4M records (+4.96%), (5) BNK_XSTAT_TX +12.3M records (+2.12%), (6) BNK_BTCH_STREF +12.3M "
            "records (+2.12%), and (7-10) ACDOCA across 4 partitions each adding approximately 9.1-9.2M records (+1.40%). "
            "The ACDOCA Universal Journal collectively accounts for over 36.6M new records, making it the highest-volume "
            "transactional entity in the system. MLDOC and MLDOCCCS exhibit above-average growth rates relative to their "
            "total sizes, indicating intense material ledger activity. These tables should be monitored for delta merge "
            "efficiency, compression ratios, and potential impact on memory and I/O performance."
        )
    ),
    (
        "CPU Spikes",
        (
            "No CPU spikes exceeding the 95% utilization threshold were detected over the last 40 days. Average CPU "
            "utilization is 3.61% across 128 logical cores (64 physical, 4 sockets), reflecting a well-balanced and "
            "healthy workload profile. No immediate corrective action is required for CPU performance. However, the "
            "long-running SQL statements identified under the P1 Alerts section should continue to be monitored, as "
            "runaway queries could cause CPU saturation during peak business processing windows if left unresolved."
        )
    ),
    (
        "Failed Backups - Type, Details, and Date",
        (
            "No failed backup executions were identified over the last 40 days. Daily full data backups are completing "
            "successfully with consistent sizes averaging approximately 2,580 GB and throughput rates ranging from 398 "
            "to 730 MB/s. Backup durations are generally stable between 58 and 108 minutes, with occasional outliers "
            "such as July 7, 2026 (190 minutes) and July 23, 2026 (107 minutes) that may indicate temporary storage I/O "
            "contention. The backup strategy is operating as expected with no coverage gaps. No remediation is required; "
            "however, continued monitoring of backup duration trends is recommended to detect early signs of storage "
            "performance degradation."
        )
    ),
]

ORANGE_COLOR = "FFA500"
HEADER_BLUE = "4472C4"
WHITE_COLOR = "FFFFFF"

num_cols = 8
last_data_row = 2 + len(rows_data)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "HANA Health Check Analysis"

ws.merge_cells("A1:" + get_column_letter(num_cols) + "1")
cap = ws["A1"]
cap.value = "DB Health check summary"
cap.font = Font(bold=True, size=14)
cap.alignment = Alignment(horizontal="center", vertical="center")
cap.fill = PatternFill(start_color=ORANGE_COLOR, end_color=ORANGE_COLOR, fill_type="solid")
ws.row_dimensions[1].height = 35

headers = [
    "Item Num", "Type", "Priority", "Item",
    "Observations", "System", "Responsible", "Link to section"
]

for col_idx, hdr in enumerate(headers, start=1):
    cell = ws.cell(row=2, column=col_idx, value=hdr)
    cell.font = Font(bold=True, color=WHITE_COLOR)
    cell.fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 30

for i, (item_text, obs_text) in enumerate(rows_data):
    row_num = i + 3

    ws.cell(row=row_num, column=1, value=i + 1)
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

    ws.cell(row=row_num, column=2, value="Database")
    ws.cell(row=row_num, column=2).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

    ws.cell(row=row_num, column=3, value="Medium")
    ws.cell(row=row_num, column=3).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

    ws.cell(row=row_num, column=4, value=item_text)
    ws.cell(row=row_num, column=4).alignment = Alignment(vertical="top", wrap_text=True)

    ws.cell(row=row_num, column=5, value=obs_text)
    ws.cell(row=row_num, column=5).alignment = Alignment(vertical="top", wrap_text=True)

    ws.cell(row=row_num, column=6, value=db_name)
    ws.cell(row=row_num, column=6).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

    ws.cell(row=row_num, column=7, value="Customer / TSM")
    ws.cell(row=row_num, column=7).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

    ws.cell(row=row_num, column=8, value="")
    ws.cell(row=row_num, column=8).alignment = Alignment(vertical="top", wrap_text=True)

    ws.row_dimensions[row_num].height = 130

ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 50
ws.column_dimensions["E"].width = 100
ws.column_dimensions["F"].width = 12
ws.column_dimensions["G"].width = 18
ws.column_dimensions["H"].width = 18

thick_side = Side(style="thick")
thin_side = Side(style="thin")

for row_idx in range(2, last_data_row + 1):
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        top = thick_side if row_idx == 2 else thin_side
        bottom = thick_side if row_idx == last_data_row else thin_side
        left = thick_side if col_idx == 1 else thin_side
        right = thick_side if col_idx == num_cols else thin_side
        cell.border = Border(top=top, bottom=bottom, left=left, right=right)

wb.save(output_path)
print("Saved:", output_path)
