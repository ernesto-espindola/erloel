import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILE_NAME = (
    r"C:\Users\I522148\OneDrive - SAP SE\SWAT\AI\Claude\working directory"
    r"\HANA_health_check_summary_observations_HP4.xlsx"
)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "HANA Health Check Analysis"

col_widths = [10, 12, 12, 60, 90, 12, 20, 18]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

thick = Side(style="thick")
medium = Side(style="medium")
thin = Side(style="thin")

caption_border = Border(left=thick, right=thick, top=thick, bottom=thick)
outer_border = Border(left=medium, right=medium, top=medium, bottom=medium)

def make_inner_border(row_idx, col_idx, total_rows, total_cols):
    left = medium if col_idx == 1 else thin
    right = medium if col_idx == total_cols else thin
    top = medium if row_idx == 1 else thin
    bottom = medium if row_idx == total_rows else thin
    return Border(left=left, right=right, top=top, bottom=bottom)

ws.merge_cells("A1:H1")
caption_cell = ws["A1"]
caption_cell.value = "DB Health check summary"
caption_cell.font = Font(name="Calibri", bold=True, size=14, color="000000")
caption_cell.fill = PatternFill(fill_type="solid", fgColor="FFA500")
caption_cell.alignment = Alignment(horizontal="center", vertical="center")
caption_cell.border = caption_border
ws.row_dimensions[1].height = 28

headers = [
    "Item Num",
    "Type",
    "Priority",
    "Item",
    "Observations",
    "System",
    "Responsible",
    "Link to section"
]

for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=2, column=col_idx, value=header)
    cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    cell.fill = PatternFill(fill_type="solid", fgColor="1F497D")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = make_inner_border(1, col_idx, 1, 8)
ws.row_dimensions[2].height = 22

obs1 = (
    "The HP4 system is running SAP HANA 2.0 SPS08 Revision 089 (version 2.00.089.00), "
    "which is the current ECS standard release. No upgrade action is required at this time. "
    "It is recommended to monitor future revision releases from SAP and plan upgrades within "
    "the standard ECS patching cycle to remain within the supported maintenance window "
    "extending through November 2028."
)

obs2 = (
    "The parameter compliance check identified 2 ERROR findings. Both relate to the "
    "sslminprotocolversion parameter being explicitly set to tls12 in global.ini under sections "
    "[communication] and [ldap], while the ECS standard expects this parameter to remain unset "
    "(EMPTY). Although TLS 1.2 enforcement is a security best practice, an explicit non-default "
    "value deviates from the ECS standard baseline and may conflict with automated configuration "
    "management. It is recommended to formally document these deviations as approved exceptions "
    "or to unset the parameters and rely on the system default TLS version enforcement. All "
    "remaining parameters are compliant with ECS standards."
)

obs3 = (
    "The PS4 tenant recorded 304 P1 alerts over the last 40 days while SYSTEMDB recorded none. "
    "The dominant issue is Alert 29 (Delta Storage Size) with 262 events, all related to RSAU_LOG "
    "(Security Audit Log), where delta storage consistently exceeded threshold with sizes ranging "
    "from 1,378 MB to 2,296 MB, indicating a persistent delta merge backlog caused by high write "
    "volume. Alert 140 (License Usage Type) generated 40 events indicating the production license "
    "usage type is not correctly configured. Alert 39 (Long-Running Statement) generated 2 events "
    "for a single SQL statement (hash 203f26e3c3018684b236ce8bcae9c2df) running between 2,005 and "
    "2,307 seconds. Recommended actions: archive or partition RSAU_LOG to reduce delta pressure, "
    "correct license usage type configuration, and analyze the long-running statement via "
    "M_SQL_PLAN_CACHE for optimization or apply a StatementTimeout."
)

obs4 = (
    "The PS4 tenant contains 17 technical tables consuming 28.85 GB of memory and 188.19 GB of "
    "disk. The five largest by disk are: SOFFCONT1 (64.33 GB, BDS document content, LOB-heavy), "
    "REPOLOAD (35.52 GB, ABAP compiled loads), RSAU_LOG (33.90 GB, Security Audit Log, 343 million "
    "rows), REPOSRC (12.51 GB, ABAP source code), and EDID4 (10.38 GB, IDoc data, 57.8 million "
    "rows). SOFFCONT1 and REPOLOAD should be reviewed for ArchiveLink or nearline storage archiving. "
    "RSAU_LOG requires urgent archiving via SM18 or SARA given active delta merge alerts. EDID4 and "
    "CDPOS are archivable via SARA standard archiving objects. A monthly review of M_CS_TABLES is "
    "recommended to track growth trends."
)

obs5 = (
    "No tables or partitions exceeding 1.5 billion records were identified in either SYSTEMDB or "
    "the PS4 tenant, which is a positive finding. However, RSAU_LOG with 343 million rows and "
    "D010TAB with 146 million rows represent tables with significant record counts that should be "
    "monitored monthly via M_CS_TABLES and prioritized for archiving before partitioning thresholds "
    "become a concern."
)

obs6 = (
    "No OOM events were recorded in either SYSTEMDB or the PS4 tenant over the last 40 days, "
    "which is a positive finding. Current memory utilisation shows 157.56 GB used out of 194.80 GB "
    "allocated (80.9% utilisation) on a 251 GB physical server. The top memory consumers in the "
    "indexserver are the ColumnStore Dictionary (26.89 GB, 13.91%), CS Buffer Page pool (21.53 GB, "
    "11.14%), and the UnifiedTableContainer (12.03 GB, 5.2%). While no OOM events occurred, the "
    "80.9% allocation utilisation warrants proactive management. Preventive recommendations include "
    "reviewing the global_allocation_limit setting, prioritising archiving of SOFFCONT1 and RSAU_LOG "
    "to reduce memory pressure, and establishing a weekly monitoring routine via "
    "M_DEV_MEMORY_COMPONENT_ALLOCATORS."
)

obs7 = (
    "A dedicated top-10 DML transaction activity query (M_TABLE_STATISTICS with "
    "WRITE_COUNT/UPDATE_COUNT filters) was not included in this health check script execution. "
    "Based on available data, tables with highest write activity indicators are: RSAU_LOG (343 "
    "million rows, active delta merge alerts from continuous security audit inserts), EDID4 (57.8 "
    "million rows, IDoc processing), CDPOS (33.2 million rows, change document writes), SWWLOGHIST "
    "(25.2 million rows, workflow processing), and DBTABLOG (9.8 million rows, table change logging). "
    "It is recommended to run a targeted M_CS_TABLES query filtering on WRITE_COUNT and "
    "DELTA_MERGE_COUNT to quantify DML activity and identify archiving priorities."
)

obs8 = (
    "One CPU spike at or above the 95% threshold was recorded over the last 40 days, on 2026-06-24 "
    "at 09:16:14 with CPU at 95%. The 40-day average CPU utilisation is 0.62%, indicating normal "
    "operation and an isolated spike event. Cross-referencing with Alert 39 long-running statement "
    "activity (hash 203f26e3c3018684b236ce8bcae9c2df detected in P1 alerts) suggests the spike may "
    "be correlated with a resource-intensive query. Recommendations: correlate the spike timestamp "
    "against M_LOAD_HISTORY_SERVICE, review the long-running SQL hash in M_SQL_PLAN_CACHE for "
    "optimization, and evaluate implementing SAP HANA Workload Classes and a StatementTimeout parameter."
)

obs9 = (
    "A total of 24 backup failures were recorded, all concentrated on a single date: June 8, 2026. "
    "Failures occurred across 8 consecutive hourly cycles from 02:00 to 09:00. Both PS4 (complete "
    "data backup, approximately 11 seconds before failure) and SYSTEMDB (immediate rejection within "
    "1 second with rapid retry) were affected in each cycle, indicating a systemic Backint or Azure "
    "storage connectivity issue rather than a data-level problem. The SYSTEMDB double-failure pattern "
    "per cycle confirms an active automated retry mechanism. Recommended actions: confirm a successful "
    "backup exists after June 8 2026; investigate Backint agent logs and Azure storage connectivity "
    "logs for that date; review and correct backup retry configuration to prevent retry storms; and "
    "implement backup monitoring alerts via M_BACKUP_CATALOG to detect failure patterns in near-real-time."
)

items = [
    "HANA Database Version",
    "ECS Standard Parameter Compliance",
    "P1 Alerts - Last 40 Days",
    "Big Technical Tables - Memory and Disk Consumption",
    "Tables or Partitions Exceeding 1.5 Billion Records",
    "Out-of-Memory (OOM) Events",
    "Top 10 Tables with Highest Transaction Activity",
    "CPU Spikes",
    "Failed Backups"
]

observations = [obs1, obs2, obs3, obs4, obs5, obs6, obs7, obs8, obs9]

color_odd = "FFFFFF"
color_even = "EBF3FB"

centered_cols = {1, 2, 3, 6, 7}

for row_num in range(3, 12):
    data_row_idx = row_num - 3
    item_num = data_row_idx + 1
    bg_color = color_odd if item_num % 2 != 0 else color_even
    fill = PatternFill(fill_type="solid", fgColor=bg_color)

    row_data = [
        item_num,
        "Database",
        "Medium",
        items[data_row_idx],
        observations[data_row_idx],
        "HP4",
        "Customer / TSM",
        ""
    ]

    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = Font(name="Calibri", size=10)
        cell.fill = fill
        if col_idx in centered_cols:
            cell.alignment = Alignment(
                horizontal="center", vertical="top", wrap_text=True
            )
        else:
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
        cell.border = make_inner_border(
            row_num - 1, col_idx, 10, 8
        )

    ws.row_dimensions[row_num].height = 120

ws.freeze_panes = "A3"

wb.save(FILE_NAME)
print(f"File saved: {FILE_NAME}")
